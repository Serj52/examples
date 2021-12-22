from sqlalchemy import create_engine
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy import Column, Integer, String
from sqlalchemy.schema import CreateTable
from sqlalchemy.orm import sessionmaker
import openpyxl
import logging

class Myclass:
    Base = declarative_base()
    class Student(Base):
        __tablename__ = 'Student'
        student_id = Column(Integer, primary_key=True)
        surname = Column(String)
        name = Column(String)
        univ_id = Column(Integer, primary_key=True)
        def __repr__(self):
            return "<'Student(student_id='{}', surname='{}',  name='{}, univ_id = '{}')>".format(
                    self.student_id, self.surname, self.name, self.univ_id)

    class University(Base):
        __tablename__ = 'University'
        univ_id = Column(Integer, primary_key=True)
        univ_name = Column(String)
        rating = (Integer)
        def __repr__(self):
            return "<'University(univ_id='{}', univ_name='{}',  rating='{}')>".format(
                    self.univ_id, self.univ_name, self.rating)

    class Subject(Base):
        __tablename__ = 'Subject'
        subj_id = Column(Integer, primary_key=True)
        subj_name = Column(String)
        def __repr__(self):
            return "<'Subject(subj_id='{}', subj_name='{}')>".format(
                    self.subj_id, self.subj_name)

    class Exam_marks(Base):
        __tablename__ = 'Exam_marks'
        student_id = Column(Integer, primary_key=True)
        mark = Column(Integer)
        subj_id = Column(Integer, primary_key=True)
        def __repr__(self):
            return "<'Exam_marks(student_id='{}', student_id='{}', subj_id = '{}')>".format(
                    self.student_id, self.student_id, self.Subj_id)

class creat_table(Myclass):
    def __init__(self, name_table, name_base = 'Stady_Base.db'):
        self.name_base = name_base
        self.name_table = name_table
        if name_table == 'Student':
            self.res = self.Student()
        elif name_table == 'University':
            self.res = self.University()
        elif name_table == 'Subject':
            self.res = self.Subject()
        elif name_table == 'Exam_marks':
            self.res = self.Exam_marks()
        self.engine = create_engine(r'sqlite:///C:\Users\Анна\PycharmProjects\untitled1\Project\{}'.format(name_base))
        self.Base.metadata.create_all(self.engine)
        print (CreateTable(self.res.__table__).compile(self.engine))

class Session_maker(Myclass):
        def __init__(self, name_base = 'Stady_Base.db'):
            self.engine = create_engine(r'sqlite:///C:\Users\Анна\PycharmProjects\untitled1\Project\{}'.format(name_base))
            Session = sessionmaker(bind=self.engine)
            self.session = Session()
            print('Сессия открыта')

        def session_add(self, db_object):
            self.session.add(db_object)
            self.session.commit()
            print('Объект {} добавлен'.format(self))

class creat_object_base(Myclass):
    def __init__(self, name_table):
        self.name = name_table
        wb = openpyxl.load_workbook(self.name)
        self.sheet = wb.active
        self.rows_max = self.sheet.max_row
        self.cols_max = self.sheet.max_column

    def add_object(self):
        for i in range(2, self.rows_max + 1):
            self.object = Myclass.Exam_marks()
            for y in range(1, self.cols_max + 1):
                res = self.sheet.cell(row = 1, column = y).value
                val = self.sheet.cell(row = i, column = y).value
                if res == 'student_id':
                    self.object.student_id = val
                elif res == 'mark':
                    self.object.mark = val
                elif res == 'subj_id':
                    self.object.subj_id = val
            Session_maker().session_add(self.object)
        logging.basicConfig(filename="journal.log",  format='%(asctime)s : %(message)s', level=logging.INFO)
        logging.info('Записи выгружены')


if __name__ == '__main__':
    g= creat_object_base('Exam_marks.xlsx').add_object()
